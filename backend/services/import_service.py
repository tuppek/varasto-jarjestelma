from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from models import Product
from services.inventory import create_product


PRODUCT_HEADERS = {
    "sku": ["sku", "tuotekoodi", "koodi", "artikkeli"],
    "name": ["name", "nimi", "tuotenimi", "tuote"],
    "description": ["description", "kuvaus"],
    "unit": ["unit", "yksikkö", "yksikko"],
    "quantity_on_hand": ["quantity_on_hand", "saldo", "varasto", "määrä", "maara", "qty"],
    "quantity_ordered": ["quantity_ordered", "tilattu", "tilattu_saldo"],
    "quantity_reserved": ["quantity_reserved", "varattu", "varattu_saldo"],
    "min_stock_level": ["min_stock_level", "minimi", "min_varasto", "hälytysraja", "halytysraja"],
    "manufacturer": ["manufacturer", "valmistaja", "brand", "merkki"],
    "wholesaler": ["wholesaler", "tukkuri", "supplier", "toimittaja", "jakelija"],
    "purchase_price": ["purchase_price", "ostohinta", "osto_hinta", "ostohinta_eur"],
    "sale_price": ["sale_price", "myyntihinta", "myynti_hinta", "myyntihinta_eur", "hinta"],
}


def _normalize(value) -> str:
    return str(value or "").strip().lower()


def _find_column(headers: list, aliases: List[str]) -> Optional[int]:
    normalized = [_normalize(h) for h in headers]
    for alias in aliases:
        if alias in normalized:
            return normalized.index(alias)
    return None


def import_products_from_excel(db: Session, file_bytes: bytes) -> dict:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"created": 0, "updated": 0, "errors": ["Tiedosto on tyhjä tai ilman dataa"]}

    headers = [str(h or "") for h in rows[0]]
    col_map = {}
    for field, aliases in PRODUCT_HEADERS.items():
        idx = _find_column(headers, aliases)
        if idx is not None:
            col_map[field] = idx

    if "sku" not in col_map or "name" not in col_map:
        return {
            "created": 0,
            "updated": 0,
            "errors": ["Pakolliset sarakkeet puuttuvat: SKU ja Nimi (tai vastaavat)"],
        }

    created = 0
    updated = 0
    errors: List[str] = []

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        try:
            sku = str(row[col_map["sku"]]).strip()
            name = str(row[col_map["name"]]).strip()
            if not sku or not name:
                errors.append(f"Rivi {row_num}: SKU ja nimi ovat pakollisia")
                continue

            data = {"sku": sku, "name": name}
            for field in (
                "description",
                "unit",
                "quantity_on_hand",
                "quantity_ordered",
                "quantity_reserved",
                "min_stock_level",
                "manufacturer",
                "wholesaler",
                "purchase_price",
                "sale_price",
            ):
                if field in col_map:
                    raw = row[col_map[field]]
                    if raw is None or str(raw).strip() == "":
                        continue
                    if field in ("quantity_on_hand", "quantity_ordered", "quantity_reserved", "min_stock_level"):
                        data[field] = int(float(raw))
                    elif field in ("purchase_price", "sale_price"):
                        data[field] = float(raw)
                    else:
                        data[field] = str(raw).strip()

            existing = db.query(Product).filter(Product.sku == sku).first()
            if existing:
                for key, value in data.items():
                    if key != "sku":
                        setattr(existing, key, value)
                updated += 1
            else:
                create_product(db, data)
                created += 1
        except Exception as exc:
            errors.append(f"Rivi {row_num}: {exc}")

    return {"created": created, "updated": updated, "errors": errors}


def create_sample_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tuotteet"
    ws.append(
        [
            "SKU",
            "Nimi",
            "Kuvaus",
            "Yksikkö",
            "Saldo",
            "Tilattu",
            "Varattu",
            "Minimi",
            "Valmistaja",
            "Tukkuri",
            "Ostohinta",
            "Myyntihinta",
        ]
    )
    ws.append(["SKU-001", "Ruuvi M6", "Teräsruuvi", "kpl", 100, 0, 0, 20, "FixPlus", "Rautakauppa Oy", 0.15, 0.35])
    ws.append(["SKU-002", "Mutteri M6", "Teräsmutteri", "kpl", 80, 50, 10, 15, "FixPlus", "Rautakauppa Oy", 0.12, 0.29])
    ws.append(["SKU-003", "Kiinnike", "Seinäkiinnike", "kpl", 25, 100, 5, 10, "BuildPro", "Tukkuri Nord", 2.5, 4.9])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
